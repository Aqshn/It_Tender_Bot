import json, re, sys, argparse
from pathlib import Path
from urllib.parse import urljoin
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

BASE_URL='https://lalafo.az'
ROOT=Path(__file__).resolve().parents[2]
input_file = ROOT/'all_lalafo_checked_v2.json'

parser = argparse.ArgumentParser(description='Run headed Playwright recheck for sellers')
parser.add_argument('--start', type=int, default=1, help='start index (1-based)')
parser.add_argument('--end', type=int, default=0, help='end index (inclusive); 0 means last')
parser.add_argument('--save-json', default='all_lalafo_checked_final.json', help='output json filename')
parser.add_argument('--save-xlsx', default='all_lalafo_checked_clean_final.xlsx', help='output xlsx filename')
args = parser.parse_args()

if not input_file.exists():
    print('Missing input:', input_file)
    sys.exit(1)

data = json.loads(input_file.read_text(encoding='utf-8'))
sellers = data.get('sellers', [])

start = max(1, args.start)
end = args.end or len(sellers)
if end > len(sellers): end = len(sellers)

print(f'Will check sellers {start}..{end} ({len(sellers)} total)')

def norm(u):
    if not u: return ''
    return urljoin(BASE_URL, u)

def scroll(page):
    last = page.evaluate('() => document.body.scrollHeight')
    for _ in range(80):
        page.evaluate('() => window.scrollTo(0, document.body.scrollHeight)')
        page.wait_for_timeout(500)
        new = page.evaluate('() => document.body.scrollHeight')
        if new==last:
            break
        last=new

mismatches=[]
updated=0
checked=0

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    ctx = browser.new_context(viewport={'width':1280,'height':900})
    page = ctx.new_page()

    for i in range(start, end+1):
        idx = i-1
        s = sellers[idx]
        shop_url = (s.get('shop_url') or '').strip()
        sid = str(s.get('seller_id') or '').strip()
        if shop_url and shop_url.startswith('http'):
            url = shop_url
        elif sid and sid.isdigit():
            url = f'{BASE_URL}/user/{sid}'
        elif shop_url:
            url = norm(shop_url)
        else:
            url = f'{BASE_URL}/user/{sid}' if sid else ''
        if not url:
            print(f'[{i}/{len(sellers)}] no url; skipping')
            continue
        try:
            print(f'[{i}/{len(sellers)}] Visiting {url}')
            resp = page.goto(url, timeout=60000)
            if resp and resp.status == 404:
                print(' -> 404'); continue
            page.wait_for_timeout(1000)
            scroll(page)
            page.wait_for_timeout(500)
            html = page.content() or ''
            anchors = set()
            for a in page.query_selector_all("a[href*='/ads/']"):
                try:
                    h = (a.get_attribute('href') or '').strip()
                    if h:
                        anchors.add(norm(h).split('#')[0])
                except Exception:
                    pass
            regex_links = set(re.findall(r"/ads/[^\"'\s]+", html))
            regex_links = set([norm(x) for x in regex_links])
            combined = anchors.union(regex_links)
            combined = [c for c in combined if '/ads/' in c]
            count = len(set(combined))
            checked += 1
            old = int(s.get('ads_count') or 0)
            if count != old:
                mismatches.append({'seller_id': s.get('seller_id'), 'shop_name': s.get('shop_name'), 'url': url, 'old': old, 'new': count})
                s['ads_count'] = count
                updated += 1
                print(f'  -> updated {s.get("seller_id")} {s.get("shop_name")} {count}')
            else:
                print(f'  -> ok {count}')
            # small pause so you can watch the page
            page.wait_for_timeout(400)
        except Exception as e:
            print('  -> error', e)
            continue

    browser.close()

# write outputs
out_json = ROOT/args.save_json
data['sellers'] = sellers
data['total_pro_sellers'] = len(sellers)
out_json.write_text(json.dumps(data, ensure_ascii=False, indent=4), encoding='utf-8')

# write excel
wb = Workbook(); ws = wb.active; ws.title = 'Lalafo'
max_mobiles = min(max((len(s.get('mobile_numbers',[])) for s in sellers), default=0), 6)
headers = ['seller_id','shop_name','shop_url','ads_count'] + [f'mobile_{i+1}' for i in range(max_mobiles)] + ['mobile_all','categories','categories_count']
ws.append(headers)
for s in sorted(sellers, key=lambda x:(-int(x.get('ads_count',0) or 0),(x.get('shop_name') or '').lower())):
    row = [s.get('seller_id',''), s.get('shop_name',''), s.get('shop_url',''), s.get('ads_count',0)]
    mobiles = s.get('mobile_numbers',[])[:max_mobiles]
    mobiles += ['']*(max_mobiles - len(mobiles))
    row += mobiles
    row.append(';'.join(s.get('mobile_numbers',[])))
    cats = s.get('categories',[])
    row.append(';'.join(cats))
    row.append(len(cats))
    ws.append(row)
for r in range(2, 2+len(sellers)):
    cell = ws.cell(row=r, column=3)
    if cell.value:
        cell.hyperlink = cell.value; cell.style = 'Hyperlink'
for cell in ws[1]:
    cell.font = Font(bold=True); cell.alignment = Alignment(wrapText=True, vertical='center')
ws.freeze_panes = 'A2'
for col in ws.columns:
    max_len = 0; col_letter = col[0].column_letter
    for cell in col:
        v = cell.value or ''
        s = str(v)
        if len(s) > max_len: max_len = len(s)
    ws.column_dimensions[col_letter].width = min(max(12, int(max_len*1.1)), 80)

out_xlsx = ROOT/args.save_xlsx
wb.save(out_xlsx)

print('\nFINAL done checked=', checked, 'updated=', updated)
print('Mismatches sample:', mismatches[:50])
